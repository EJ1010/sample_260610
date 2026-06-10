import streamlit as st
import pandas as pd
from datetime import datetime
import os
from openpyxl import load_workbook, Workbook

# 1. 페이지 기본 설정 및 제목
st.set_page_config(page_title="Streamlit 카페", layout="centered")
st.title('☕ 미니 카페 주문 시스템')
st.subheader('KDT 파이썬 미션 : 영수증 발행 및 엑셀 저장')

# 엑셀 파일 저장 경로 설정 (현재 프로그램이 있는 폴더에 저장)
EXCEL_PATH = '카페주문내역.xlsx'

# 2. Session State 초기화 (주문 내역을 저장할 장소)
if 'order_list' not in st.session_state:
    st.session_state.order_list = []

# 메뉴판 데이터
MENU = {
    '아메리카노': 4000,
    '카페라떼': 4500,
    '바닐라라떼': 5000,
    '그린티 라떼': 5500,
    '치즈케이크': 6000,
    '크로플': 4500
}

# 3. 사이드바 구성
st.sidebar.header("⚙️ 포스기 관리")
menu_choice = st.sidebar.selectbox("기능 선택", ['주문하기', '전체 매출 확인(준비중)'])

# 4. 메인 화면 탭 구성
tab1, tab2 = st.tabs(['🛒 메뉴 주문', '🧾 영수증 출력'])

# --- Tab 1: 메뉴 주문 ---
with tab1:
    st.markdown("### 📣 원하시는 메뉴를 선택해주세요.")
    
    with st.form('order_form'):
        c1, c2 = st.columns(2)
        item = c1.selectbox('메뉴 선택', list(MENU.keys()))
        quantity = c2.number_input('수량', min_value=1, max_value=10, value=1, step=1)
        
        c3, c4 = st.columns(2)
        cup_type = c3.radio('컵 선택', ['매장용 컵', '일회용 컵'])
        options = c4.multiselect('추가 옵션', ['샷 추가(+500원)', '시럽 추가(+500원)', '얼음 많이'])
        
        submit_button = st.form_submit_button('장바구니 담기')
        
        if submit_button:
            extra_price = 0
            if '샷 추가(+500원)' in options: extra_price += 500
            if '시럽 추가(+500원)' in options: extra_price += 500
            
            unit_price = MENU[item] + extra_price
            total_price = unit_price * quantity
            
            order_item = {
                '메뉴명': item,
                '단가': unit_price,
                '수량': quantity,
                '금액': total_price,
                '포장여부': cup_type,
                '옵션': ", ".join(options) if options else "없음"
            }
            
            st.session_state.order_list.append(order_item)
            st.success(f"✅ {item} {quantity}잔이 장바구니에 담겼습니다!")

    if st.session_state.order_list:
        st.markdown("---")
        st.markdown("#### 🛒 현재 장바구니 내역")
        df_current = pd.DataFrame(st.session_state.order_list)
        st.table(df_current[['메뉴명', '단가', '수량', '금액']])
        
        if st.button('🛒 장바구니 비우기'):
            st.session_state.order_list = []
            st.rerun()


# --- Tab 2: 영수증 출력 및 엑셀 저장 ---
with tab2:
    st.markdown("### 🧾 주문 영수증")
    
    if not st.session_state.order_list:
        st.warning("주문 내역이 없습니다. 먼저 메뉴를 주문해주세요!")
    else:
        current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        # 영수증 디자인 박스
        with st.container(border=True):
            st.markdown("<h2 style='text-align: center;'>📄 RECEIPT 📄</h2>", unsafe_allow_html=True)
            st.write(f"**상호명:** KDT 카페 스트림릿점")
            st.write(f"**일시:** {current_time}")
            st.markdown("---")
            
            grand_total = 0
            for idx, order in enumerate(st.session_state.order_list, 1):
                st.write(f"**{idx}. {order['메뉴명']}** ({order['포장여부']})")
                st.caption(f"↳ 옵션: {order['옵션']}")
                st.write(f"   {order['단가']:,}원 × {order['수량']}개 = **{order['금액']:,}원**")
                grand_total += order['금액']
                
            st.markdown("---")
            
            vat = int(grand_total * 0.1)
            supply_value = grand_total - vat
            
            c1, c2 = st.columns(2)
            c1.write("가액:")
            c2.write(f"{supply_value:,} 원")
            c1.write("부가세(10%):")
            c2.write(f"{vat:,} 원")
            
            st.markdown("### **총 결제금액: " + f"{grand_total:,} 원**")
            st.markdown("<p style='text-align: center; color: gray;'>이용해 주셔서 감사합니다!</p>", unsafe_allow_html=True)
            
        # 💳 결제 완료 버튼 클릭 시 엑셀 파일에 저장
        if st.button('💳 결제 완료 (엑셀 저장)'):
            
            # 1. 엑셀 파일이 없으면 새로 만들고 헤더 추가, 있으면 불러오기
            if not os.path.exists(EXCEL_PATH):
                wb = Workbook()
                ws = wb.active
                ws.title = "주문내역"
                # 헤더 작성
                ws.append(["주문일시", "메뉴명", "단가", "수량", "총금액", "포장여부", "선택옵션"])
            else:
                wb = load_workbook(EXCEL_PATH)
                ws = wb.active
                
            # 2. 장바구니(order_list)에 있는 항목들을 행(Row)으로 추가
            for order in st.session_state.order_list:
                ws.append([
                    current_time,
                    order['메뉴명'],
                    order['단가'],
                    order['수량'],
                    order['금액'],
                    order['포장여부'],
                    order['옵션']
                ])
            
            # 3. 엑셀 파일 저장 및 종료
            wb.save(EXCEL_PATH)
            wb.close()
            
            # 4. 완료 처리 및 세션 초기화
            st.session_state.order_list = []
            st.balloons()
            st.success(f"🎉 결제가 완료되었으며, 내역이 '{EXCEL_PATH}' 파일에 저장되었습니다!")
            st.rerun()